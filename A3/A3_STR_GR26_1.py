import ifcopenshell
import ifcopenshell.util.element
import openpyxl
import ifcopenshell.geom
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Load the IFC model of the building module from a file
model = ifcopenshell.open('C:/Users/friis/OneDrive - Danmarks Tekniske Universitet/DTU/41934 - Advanced BIM/IFC/CES_BLD_24_06_STR.ifc')
beams = model.by_type('IfcBeam')  # Retrieve all beam elements from the model

# Extract information about each beam, including type, GlobalId, and span length
beam_types = [beam.ObjectType for beam in beams]
beam_globalId = [beam.GlobalId for beam in beams]
beam_spans = [
    next((prop.NominalValue.wrappedValue / 1000 for rel in beam.IsDefinedBy for prop in rel.RelatingPropertyDefinition.HasProperties if prop.Name == 'Span'), 0)
    for beam in beams]

# Calculate the quantity and total span length for each beam type
type_quantity = {name: beam_types.count(name) for name in set(beam_types)}
type_spans = {}
for i, name in enumerate(beam_types):
    if name not in type_spans:
        type_spans[name] = 0
    span = beam_spans[i]
    if span:
        type_spans[name] += span

# Compute the total span of all beams in the model
total_span = sum(prop.NominalValue.wrappedValue for entity in beams for rel in entity.IsDefinedBy for prop in rel.RelatingPropertyDefinition.HasProperties if prop.Name == 'Span')

# Function to calculate the dimensions (length, width, height) of any given entity
def calculate_entity_dimensions(element):
    config = ifcopenshell.geom.settings()
    config.set(config.USE_WORLD_COORDS, True)
    geometry = ifcopenshell.geom.create_shape(config, element)
    coordinates = geometry.geometry.verts
    x, y, z = max(coordinates[0::3]) - min(coordinates[0::3]), max(coordinates[1::3]) - min(coordinates[1::3]), max(coordinates[2::3]) - min(coordinates[2::3])
    return x, y, z

# Function to gather data on dimensions and volume/length for walls, columns, or slabs
def gather_entity_data(entity_type, entity_summary, data_matrix, measure='volume'):
    entities = model.by_type(entity_type)
    for element in entities:
        lx, ly, lz = calculate_entity_dimensions(element)
        measure_value = lx * ly * lz if measure == 'volume' else lz  # Calculate volume or length based on measure type
        data_matrix.append([element.GlobalId, element.ObjectType, measure_value])
        entity_summary[element.ObjectType] = entity_summary.get(element.ObjectType, 0) + measure_value

# Initialize matrices and summaries for walls, columns, and slabs
walls_data, columns_data, slabs_data = [], [], []
walls_summary, columns_summary, slabs_summary = {}, {}, {}

# Gather data for each entity type (wall, column, slab)
gather_entity_data('IfcWall', walls_summary, walls_data)
gather_entity_data('IfcColumn', columns_summary, columns_data, measure='length')
gather_entity_data('IfcSlab', slabs_summary, slabs_data)

# Convert summary dictionaries to data matrices for easy export
def convert_summary_to_matrix(summary_dict, entity_list):
    return [[entity_type, sum(1 for element in entity_list if element.ObjectType == entity_type), total_value] for entity_type, total_value in summary_dict.items()]

# Create matrices of beam, wall, column, and slab data to be inserted into Excel
beam_data_matrix = [[beam_type, type_quantity[beam_type], type_spans[beam_type]] for beam_type in type_spans]
wall_data_matrix = convert_summary_to_matrix(walls_summary, model.by_type('IfcWall'))
column_data_matrix = convert_summary_to_matrix(columns_summary, model.by_type('IfcColumn'))
slab_data_matrix = convert_summary_to_matrix(slabs_summary, model.by_type('IfcSlab'))

# Initialize a new Excel workbook and remove the default sheet
workbook = openpyxl.Workbook()
workbook.remove(workbook.active)

# Create separate sheets for different types of data
sheet_titles = ["Overview", "Beams", "Columns", "Walls", "Slabs"]
for title in sheet_titles:
    workbook.create_sheet(title)

# Access each sheet by name
overview_sheet = workbook["Overview"]
beams_sheet = workbook["Beams"]
columns_sheet = workbook["Columns"]
walls_sheet = workbook["Walls"]
slabs_sheet = workbook["Slabs"]

# Set up the layout for the Overview sheet
overview_sheet.cell(row=3, column=1, value="Beam")
overview_sheet.cell(row=2, column=2, value="Estimated Price pr. element[kr]")
overview_sheet.cell(row=4, column=1, value="Column")
overview_sheet.cell(row=5, column=1, value="Wall")
overview_sheet.cell(row=6, column=1, value="Slab")
overview_sheet.cell(row=7, column=1, value="Total price for all structural elements")

# Function to insert headers and data rows into a specific sheet
def insert_data_into_sheet(sheet, headers, data_matrix):
    #Inserts headers and data into a specified sheet
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)
    for row_num, data_row in enumerate(data_matrix, start=2):
        for col_num, value in enumerate(data_row, start=1):
            sheet.cell(row=row_num, column=col_num, value=value)

# Define headers and insert data into each sheet (beams, columns, walls, slabs)
beam_headers = ["Beam Type", "Quantity", "Total Length [m]", "Estimated Cost pr. Length [kr./m]", "Computed Cost [kr.]"]
insert_data_into_sheet(beams_sheet, beam_headers, beam_data_matrix)

column_headers = ["Column Type", "Quantity", "Total Volume [m³]", "Estimated Cost pr. Volume [kr./m³]", "Computed Cost [kr.]"]
insert_data_into_sheet(columns_sheet, column_headers, column_data_matrix)

wall_headers = ["Wall Type", "Quantity", "Total Volume [m³]", "Estimated Cost pr. Volume [kr./m³]", "Computed Cost [kr.]"]
insert_data_into_sheet(walls_sheet, wall_headers, wall_data_matrix)

slab_headers = ["Slab Type", "Quantity", "Total Volume [m³]", "Estimated Cost pr. Volume [kr./m³]", "Computed Cost [kr.]"]
insert_data_into_sheet(slabs_sheet, slab_headers, slab_data_matrix)

# Save the workbook to a specified file path
workbook.save(r'C:/Users/friis/OneDrive - Danmarks Tekniske Universitet/DTU/41934 - Advanced BIM/A3/StucturalElementDefinetions_AddPrices.xlsx')

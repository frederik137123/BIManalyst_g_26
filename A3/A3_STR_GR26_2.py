import ifcopenshell
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# Load existing Excel file from previous script with inserted prices
price_data_path = r"C:/Users/friis/OneDrive - Danmarks Tekniske Universitet/DTU/41934 - Advanced BIM/A3/StucturalElementDefinetions_AddPrices.xlsx"
workbook = openpyxl.load_workbook(price_data_path)
overview_sheet = workbook["Overview"]  # Get the Overview sheet for totals


beams_sheet = workbook["Beams"]
columns_sheet = workbook["Columns"]
walls_sheet = workbook["Walls"]
slabs_sheet = workbook["Slabs"]

# Create a dictionary to easily iterate through each sheet
sheet_map = {
    'Beams': beams_sheet,
    'Columns': columns_sheet,
    'Walls': walls_sheet,
    'Slabs': slabs_sheet
}

# Initialize variables for the calculations
total_price = 0
row_count = 3  # Starting row in Overview sheet for each component's total
slabs_adjusted_total = 0  # Variable to hold slab total adjusted with a specific beam type price

# Loop through each sheet to calculate total prices for beams, columns, walls, and slabs
for sheet_name, worksheet in sheet_map.items():
    total = 0  # Initialize total price for the current sheet
    
    # Loop through each row to calculate the price for each item
    for i in range(2, worksheet.max_row + 1):
        total_length_or_volume = worksheet.cell(row=i, column=3).value  # Column C
        estimated_price_per_unit = worksheet.cell(row=i, column=4).value  # Column D
        
        # Calculate price if both values are present, otherwise set to 0
        if total_length_or_volume is not None and estimated_price_per_unit is not None:
            calculated_price = total_length_or_volume * estimated_price_per_unit
        else:
            calculated_price = 0  # Default to 0 if missing data

        # Check if the current sheet is "Beams" and if this beam type's price should be added to slabs
        if sheet_name == 'Beams':
            beam_type = worksheet.cell(row=i, column=1).value  # Assuming Column B holds the beam type
            if beam_type == 'EX18 Precast-Hollow Core Slab_Spæncom:STR - Hollow Core Slab - EX18':  # Replace with the actual beam type to count toward slabs
                slabs_adjusted_total += calculated_price  # Add to slabs total instead
                continue  # Skip adding this price to beams total below
        
        # Write the calculated price to Column E
        worksheet.cell(row=i, column=5, value=calculated_price)
        total += calculated_price  # Add to total for the sheet
    
    # Write subtotal for each sheet in its last row
    worksheet.cell(row=worksheet.max_row + 2, column=5, value=total)
    overview_sheet.cell(row=row_count, column=2, value=total)  # Add subtotal to Overview sheet
    
    # Add the sheet total to grand total
    total_price += total  
    row_count += 1  # Move to the next row for the next component

# Add the adjusted slab total to slabs (including the price of the specific beam type)
overview_sheet.cell(row=row_count - 1, column=2, value=overview_sheet.cell(row=row_count - 1, column=2).value + slabs_adjusted_total)

# Write the grand total price in the Overview sheet
overview_sheet.cell(row=row_count, column=2, value=total_price)

# Add labels for each component type in the Overview sheet
overview_sheet.cell(row=3, column=1, value="Beams")
overview_sheet.cell(row=4, column=1, value="Columns")
overview_sheet.cell(row=5, column=1, value="Walls")
overview_sheet.cell(row=6, column=1, value="Slabs")

# Create a chart to represent total prices by component type
chart = BarChart()
chart.title = "Total Price by Component"
chart.x_axis.title = "Component Type"
chart.y_axis.title = "Total Price"

# Define data and labels for the chart
data = Reference(overview_sheet, min_col=2, min_row=3, max_row=row_count - 1)
labels = Reference(overview_sheet, min_col=1, min_row=3, max_row=row_count - 1)

# Add data and labels to the chart, and display data values on bars
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)
chart.dataLabels = DataLabelList(showVal=True)

# Add the chart to the Overview sheet at a specified location
overview_sheet.add_chart(chart, "E8")

# Define a function to create and add a histogram chart for each sheet
def add_histogram_chart(worksheet, title, data_start_row, data_end_row, data_col, label_col, chart_position):
    chart = BarChart()
    chart.title = title
    chart.x_axis.title = 'Type'
    chart.y_axis.title = 'Quantity'

    # Define data and labels for the histogram chart
    data = Reference(worksheet, min_col=data_col, min_row=data_start_row, max_row=data_end_row)
    labels = Reference(worksheet, min_col=label_col, min_row=data_start_row, max_row=data_end_row)

    # Add data and labels to the histogram chart
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    # Place the histogram chart in the specified worksheet location
    worksheet.add_chart(chart, chart_position)

# Add histogram charts to each component sheet for type distribution
add_histogram_chart(beams_sheet, "Beam Type Distribution", 2, beams_sheet.max_row, 2, 1, "F2")
add_histogram_chart(columns_sheet, "Column Type Distribution", 2, columns_sheet.max_row, 2, 1, "F2")
add_histogram_chart(walls_sheet, "Wall Type Distribution", 2, walls_sheet.max_row, 2, 1, "F2")
add_histogram_chart(slabs_sheet, "Slab Type Distribution", 2, slabs_sheet.max_row, 2, 1, "F2")

# Save the updated workbook with data and charts
workbook.save(r'C:/Users/friis/OneDrive - Danmarks Tekniske Universitet/DTU/41934 - Advanced BIM/A3/EstimatedPrice.xlsx')
